import re
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import csv
from fractions import Fraction

def detect_and_convert_number(value):
    currency_symbols = ['Rp', 'IDR', '$']
    currency = None
    
    for symbol in currency_symbols:
        if value.startswith(symbol):
            currency = symbol
            value = value[len(symbol):]
            break
    
    value = re.sub(r'[^\d,\.]', '', value)
    
    thousand_pattern_comma = re.compile(r'^\d{1,3}(,\d{3})*$')
    thousand_pattern_dot = re.compile(r'^\d{1,3}(\.\d{3})*$')
    
    if thousand_pattern_comma.match(value):
        value = value.replace(',', '')
        return int(value), 'comma', currency
    elif thousand_pattern_dot.match(value):
        value = value.replace('.', '')
        return int(value), 'dot', currency
    else:
        try:
            return int(value), 'none', currency
        except ValueError:
            return value, None, currency

def logika_angka_bulat(value):
    if isinstance(value, int):
        return value
    if value.is_integer():
        return int(value)
    return value

def sesuai_karakter_format_output(value, format_type, currency=None):
    value = logika_angka_bulat(value)
    
    if format_type == 'comma':
        result = f"{value:,}".replace(',', ',')
    elif format_type == 'dot':
        result = f"{value:,}".replace(',', '.')
    else:
        result = str(value)
    
    if currency:
        result = f"{currency} {result}"
    
    return result

def convert_to_numeric(value):
    if value is None:
        return None
    try:
        return float(value)
    except ValueError:
        return value

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def print_interface(content, headers, data):
    clear_screen()
    if content is None:
        content = ""
    border = '+' + '-' * 78 + '+'
    empty_line = '|' + ' ' * 78 + '|'
    lines = content.split("\n")
    centered_lines = [line.center(78) for line in lines]
    padding_top = (20 - len(centered_lines) - 3) // 2
    padding_bottom = 20 - len(centered_lines) - padding_top - 3

    print(border)
    for _ in range(padding_top):
        print(empty_line)
    print('|' + "BBYALIEN".center(78) + '|')
    print('|' + "EXCEL GENERATOR".center(78) + '|')
    print(empty_line)
    for line in centered_lines:
        print('|' + line.ljust(78) + '|')
    for _ in range(padding_bottom):
        print(empty_line)
    print(border)

    if headers:
        print_table(headers, data)

def print_table(headers, data):
    col_widths = [max(len(header), 10) for header in headers]
    for row in data:
        for i, cell in enumerate(row):
            col_widths[i] = max(col_widths[i], len(str(cell)))

    max_col_width = 25
    col_widths = [min(width, max_col_width) for width in col_widths]

    terminal_width = 78
    total_width = sum(col_widths) + (3 * len(col_widths)) + 1

    if total_width > terminal_width:
        scale_factor = terminal_width / total_width
        col_widths = [int(col_width * scale_factor) for col_width in col_widths]

    border_line = '+-' + '-+-'.join('-' * width for width in col_widths) + '-+'
    print(border_line)
    print_row(headers, col_widths)
    print(border_line)

    for row in data:
        print_row(row, col_widths)
        print(border_line)

def print_row(cells, col_widths):
    row = '|'
    for i, cell in enumerate(cells):
        cell_str = str(cell)
        if len(cell_str) > col_widths[i]:
            cell_str = cell_str[:col_widths[i] - 3] + '...'
        row += f" {cell_str.ljust(col_widths[i])} |"
    print(row)

def validate_font_input():
    while True:
        font_input = input("Enter font name and size (for example, Times New Roman 12): ").rsplit(maxsplit=1)
        if len(font_input) == 1 and font_input[0] == '0':
            return None, None
        if len(font_input) == 2:
            font_name = font_input[0]
            if re.match("^[A-Za-z0-9\-_ /,]+$", font_name):
                try:
                    font_size = int(font_input[1])
                    return font_name, font_size
                except ValueError:
                    print("Invalid input. Size must be a number.")
            else:
                print("Invalid font name. Please use only allowed characters: letters, numbers, -, _, /, and commas.")
        else:
            print("Invalid input. Please enter font name and size.")

def display_page(page_number, lang):
    pages = {
        1: {
            'en': "\nChoose your language:\n\n1. English                      2. Bahasa Indonesia\n0. Exit",
            'id': "\nPilih bahasa Anda:\n\n1. English                         2. Bahasa Indonesia\n0. Keluar"
        },
        2: {
            'en': "\nChoose the type of table:\n\n1. Simple Table                  2. Excel Table\n0. Exit",
            'id': "\nPilih jenis tabel:\n\n1. Simple Table                         2. Tabel Excel\n0. Keluar"
        },
        3: {
            'en': "\nEnter headers separated by commas:\n0. Exit",
            'id': "\nMasukkan header dipisahkan dengan koma:\n0. Keluar"
        },
        4: {
            'en': "\nEnter the number of rows you want to create:\n0. Exit",
            'id': "\nMasukkan jumlah baris yang ingin Anda buat:\n0. Keluar"
        },
        5: {
            'en': "\nEnter data for the table rows:\n0. Exit",
            'id': "\nMasukkan data untuk baris tabel:\n0. Keluar"
        },
        6: {
            'en': "\nAdd borders to Simple Table? (all/top/bottom/left/right):\n0. Skip",
            'id': "\nTambahkan border ke Tabel Sederhana? (all/atas/bawah/kiri/kanan):\n0. Lewati"
        },
        7: {
            'en': "\nEnter font name and size for header:\n0. Exit",
            'id': "\nMasukkan nama dan ukuran font untuk header:\n0. Keluar"
        },
        8: {
            'en': "\nEnter font name and size for table content:\n0. Exit",
            'id': "\nMasukkan nama dan ukuran font untuk isi tabel:\n0. Keluar"
        },
        9: {
            'en': "\nChoose page layout:\n\n1. A4                            2. Letter                          3. Legal\n0. Exit",
            'id': "\nPilih tata letak halaman:\n\n1. A4                            2. Letter                          3. Legal\n0. Keluar"
        },
        10: {
            'en': "\nEnter file name:\n0. Exit",
            'id': "\nMasukkan nama file:\n0. Keluar"
        },
        11: {
            'en': "Choose operation type:\n1. Basic Math Operations\n2. Arithmetic Operations\n3. Skip\n0. Exit",
            'id': "Pilih jenis operasi:\n1. Operasi Matematika Dasar\n2. Operasi Aritmatika\n3. Lewati\n0. Keluar"
        },
        12: {
            'en': "Specify columns for basic math operations (comma-separated):",
            'id': "Pilih kolom untuk operasi matematika dasar (dipisah koma):"
        },
        13: {
            'en': "\nChoose the basic math operation:\n\n1. Addition (+)\n2. Subtraction (-)\n3. Multiplication (*)\n4. Division (/): ",
            'id': "\nPilih operasi matematika dasar:\n\n1. Penjumlahan (+)\n2. Pengurangan (-)\n3. Perkalian (*)\n4. Pembagian (/): "
        },
        14: {
            'en': "Enter arithmetic formulas separated by commas (e.g., A2*B2, C3+D3): ",
            'id': "Masukkan rumus aritmatika yang dipisahkan oleh koma (misalnya, A2*B2, C3+D3): "
        }
    }
    return pages[page_number][lang]

def validate_headers_input(headers_input):
    headers = list(csv.reader([headers_input], delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL))[0]
    headers = [str(header) for header in headers]
    while any(header.startswith("'") for header in headers):
        print("Header tidak boleh mengandung tanda petik satu. Silakan masukkan ulang.")
        headers_input = input("Masukkan header: ")
        headers = list(csv.reader([headers_input], delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL))[0]
        headers = [str(header) for header in headers]
    return headers

def process_input_row(input_row, headers):
    try:
        input_row = re.sub(r"^'|(?<=[^\\])'", '"', input_row)
        row = list(csv.reader([input_row], delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL))[0]

        if len(row) != len(headers):
            print("The number of columns in the row does not match the number of headers.")
            return None
        
        for i, value in enumerate(row):
            if value.startswith('"') and value.endswith('"'):
                row[i] = value[1:-1]
            else:
                row[i] = value
        return row
    except Exception as e:
        print(f"Error processing row: {e}")
        return None

def auto_generate_excel():
    headers = []
    data = []
    border_option = None
    use_math_operations = False

    while True:
        content = display_page(1, 'en')
        print_interface(content, headers, data)
        lang_choice = input("Enter your choice: ")
        if lang_choice == '0':
            return
        elif lang_choice == '1':
            lang = 'en'
            break
        elif lang_choice == '2':
            lang = 'id'
            break
        else:
            print("Invalid choice. Please enter 1, 2, or 0.")

    while True:
        content = display_page(2, lang)
        print_interface(content, headers, data)
        table_type = input("Enter your choice: ")
        if table_type == '0':
            return
        if table_type in ['1', '2']:
            table_type = int(table_type)
            break

    content = display_page(3, lang)
    print_interface(content, headers, data)
    while True:
        headers_input = input("Enter headers: ")
        if headers_input == '0':
            return
        headers = validate_headers_input(headers_input)
        if headers:
            break
        else:
            print("Invalid headers. Please try again.")

    while True:
        content = display_page(4, lang)
        print_interface(content, headers, data)
        num_rows = input("Enter number of rows: ")
        if num_rows == '0':
            return
        try:
            num_rows = int(num_rows)
            if num_rows <= 900000000:
                break
            else:
                print("Please enter a number less than or equal to 900,000,000.")
        except ValueError as e:
            print(f"Invalid input. Please enter a number. Error: {e}")

    for i in range(num_rows):
        while True:
            content = display_page(5, lang)
            print_interface(content, headers, data)
            row_data = input(f"Row {i + 1}: ")
            if row_data == '0':
                return
            try:
                row = process_input_row(row_data, headers)
                if row is None:
                    continue
                if len(row) != len(headers):
                    print(f"Warning: The number of columns in row {i + 1} does not match the header count. Please try again.")
                else:
                    data.append(row)
                    print_interface("", headers, data)
                    print_row(row, [len(header) for header in headers])
                    break
            except Exception as e:
                print(f"Error processing row {i + 1}: {e}")

    wb = openpyxl.Workbook()
    ws = wb.active

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=str(header))

    for row_num, row in enumerate(data, start=2):
        for col_num, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    if table_type == 1:
        content = display_page(6, lang)
        print_interface(content, headers, data)
        border_option = input("Add borders to Simple Table? (all/top/bottom/left/right): ")
        if border_option == '0':
            border_option = None

    content = display_page(7, lang)
    print_interface(content, headers, data)
    font_header_name, font_header_size = validate_font_input()
    
    content = display_page(8, lang)
    print_interface(content, headers, data)
    font_content_name, font_content_size = validate_font_input()

    while True:
        content = display_page(11, lang)
        print_interface(content, headers, data)
        operation_choice = input("Choose operation type:\n1. Basic Math Operations\n2. Arithmetic Operations\n3. Skip\n0. Exit: ")
        if operation_choice == '0':
            return
        elif operation_choice == '1':
            use_math_operations = True
            break
        elif operation_choice == '2':
            use_math_operations = False
            break
        elif operation_choice == '3':
            content = display_page(9, lang)
            print_interface(content, headers, data)
            while True:
                page_layout = input("Enter your choice: ")
                if page_layout == '0':
                    return
                if page_layout in ['1', '2', '3']:
                    page_layout = int(page_layout)
                    break

            content = display_page(10, lang)
            print_interface(content, headers, data)
            file_name = input("Enter file name: ")
            if file_name == '0':
                return

            try:
                layout_map = {1: 9, 2: 1, 3: 5}
                ws.page_setup.paperSize = layout_map.get(page_layout, 9)

                if border_option:
                    border_style = Border(
                        top=Side(style='thin'),
                        left=Side(style='thin'),
                        bottom=Side(style='thin'),
                        right=Side(style='thin')
                    )

                    for row in ws.iter_rows(min_row=1, max_row=len(data) + 1, min_col=1, max_col=len(headers)):
                        for cell in row:
                            if border_option == 'all' or (
                                (border_option == 'top' and cell.row == 1) or
                                (border_option == 'bottom' and cell.row == len(data) + 1) or
                                (border_option == 'left' and cell.column == 1) or
                                (border_option == 'right' and cell.column == len(headers))
                            ):
                                cell.border = border_style

                if table_type == 2:
                    table = Table(displayName="Table1", ref="A1:{}".format(get_column_letter(len(headers)) + str(len(data) + 1)))
                    style = TableStyleInfo(
                        name="TableStyleMedium9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=True
                    )
                    table.tableStyleInfo = style
                    ws.add_table(table)

                if font_header_name and font_header_size:
                    for col_num in range(1, len(headers) + 1):
                        ws.cell(row=1, column=col_num).font = Font(name=font_header_name, size=font_header_size, bold=True)

                if font_content_name and font_content_size:
                    for row_num in range(2, len(data) + 1):
                        for col_num in range(1, len(headers) + 1):
                            ws.cell(row=row_num, column=col_num).font = Font(name=font_content_name, size=font_content_size)

                wb.save("{}.xlsx".format(file_name))
                print("Excel file '{}' generated successfully!".format(file_name))
            except Exception as e:
                print("Error: {}".format(e))
            return

    if use_math_operations:
        while True:
            content = display_page(12, lang)
            print_interface(content, headers, data)
            basic_math_columns_input = input("Specify columns for basic math operations (comma-separated): ")
            basic_math_columns = [col.strip() for col in basic_math_columns_input.split(',')]
            if all(col in headers for col in basic_math_columns):
                break
            else:
                print("Invalid input. Please specify at least one column that matches the headers.")

        while True:
            content = display_page(13, lang)
            print_interface(content, headers, data)
            basic_math_operation = input("Choose the basic math operation:\n\n1. Addition (+)\n2. Subtraction (-)\n3. Multiplication (*)\n4. Division (/): ")
            if basic_math_operation in ['1', '2', '3', '4']:
                break
            else:
                print("Invalid input. Please enter 1, 2, 3, or 4.")

        for col in basic_math_columns:
            col_index = headers.index(col) + 1
            col_values = [detect_and_convert_number(ws.cell(row=row_index, column=col_index).value)[0] for row_index in range(2, len(data) + 2)]
            col_values = [value for value in col_values if value is not None]

            if basic_math_operation == '1':
                result = sum(col_values)
            elif basic_math_operation == '2':
                result = col_values[0] - sum(col_values[1:])
            elif basic_math_operation == '3':
                result = col_values[0] * sum(col_values[1:])
            elif basic_math_operation == '4':
                try:
                    result = col_values[0] / sum(col_values[1:])
                except ZeroDivisionError:
                    result = None

            if result is not None:
                if isinstance(result, float) and result.is_integer():
                    result = int(result)
                elif isinstance(result, float):
                    result = round(result, 2)

            format_type, currency = detect_and_convert_number(ws.cell(row=2, column=col_index).value)[1:3]
            output = sesuai_karakter_format_output(result, format_type, currency)
            ws.cell(row=len(data) + 2, column=col_index, value=output)
            ws.cell(row=len(data) + 2, column=col_index).font = Font(bold=True)

        ws.cell(row=len(data) + 2, column=1, value="JUMLAH")

    else:  # Arithmetic Operations
        content = display_page(14, lang)
        print_interface(content, headers, data)
        while True:
            arithmetic_formulas_input = input("Enter arithmetic formulas separated by commas (e.g., A2*B2, C3+D3): ")
            if arithmetic_formulas_input:
                arithmetic_formulas = [formula.strip() for formula in arithmetic_formulas_input.split(',')]
                break

        headers.append("RESULT")
        headers = [str(header) for header in headers]  # Ensure all headers are strings

        for i, formula in enumerate(arithmetic_formulas):
            for row_index in range(2, len(data) + 2):
                formula_values = {cell.coordinate: convert_to_numeric(cell.value) for cell in ws[row_index]}
                formula_values = {k: v for k, v in formula_values.items() if v is not None}  # Ignore None values
                try:
                    formula_result = eval(formula, {'__builtins__': None}, formula_values)
                except Exception as e:
                    print(f"Error evaluating formula '{formula}': {e}")
                    formula_result = None
                ws.cell(row=row_index, column=len(headers), value=formula_result)

    # Ensure all headers are strings before saving the file
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=str(header))

    content = display_page(9, lang)
    print_interface(content, headers, data)
    while True:
        page_layout = input("Enter your choice: ")
        if page_layout == '0':
            return
        if page_layout in ['1', '2', '3']:
            page_layout = int(page_layout)
            break

    content = display_page(10, lang)
    print_interface(content, headers, data)
    file_name = input("Enter file name: ")
    if file_name == '0':
        return

    try:
        layout_map = {1: 9, 2: 1, 3: 5}
        ws.page_setup.paperSize = layout_map.get(page_layout, 9)

        if border_option:
            border_style = Border(
                top=Side(style='thin'),
                left=Side(style='thin'),
                bottom=Side(style='thin'),
                right=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=1, max_row=len(data) + 2, min_col=1, max_col=len(headers)):
                for cell in row:
                    if border_option == 'all' or (
                        (border_option == 'top' and cell.row == 1) or
                        (border_option == 'bottom' and cell.row == len(data) + 2) or
                        (border_option == 'left' and cell.column == 1) or
                        (border_option == 'right' and cell.column == len(headers))
                    ):
                        cell.border = border_style

        if table_type == 2:
            table = Table(displayName="Table1", ref="A1:{}".format(get_column_letter(len(headers)) + str(len(data) + 2)))
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        if font_header_name and font_header_size:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=1, column=col_num).font = Font(name=font_header_name, size=font_header_size, bold=True)

        if font_content_name and font_content_size:
            for row_num in range(2, len(data) + 2):
                for col_num in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col_num).font = Font(name=font_content_name, size=font_content_size)

        wb.save("{}.xlsx".format(file_name))
        print("Excel file '{}' generated successfully!".format(file_name))
    except Exception as e:
        print("Error: {}".format(e))

if __name__ == "__main__":
    auto_generate_excel()
